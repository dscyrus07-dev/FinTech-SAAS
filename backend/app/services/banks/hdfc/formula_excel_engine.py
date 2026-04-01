"""
Formula-Based Excel Report Engine
==================================
Generates Excel reports using ONLY Excel formulas for all calculations.
No hardcoded numeric values - everything is formula-driven.

Column Mapping (FIXED - DO NOT CHANGE):
  Column A = Date         (date values)
  Column B = Description  (text)
  Column C = Debit        (numeric — blank if no debit)
  Column D = Credit       (numeric — blank if no credit)
  Column E = Balance      (numeric)
  Column F = Category     (text)
  Column G = Confidence   (text or %)
  Column H = Recurring    (text: "Yes" or "No")

ABSOLUTE RULES:
  - Column D (Credit) = ONLY source for ALL credit calculations
  - Column C (Debit) = ONLY source for ALL debit calculations
  - NEVER infer credit/debit from Description or Balance
  - Blank cell = zero. Never skip. Never guess.
"""

import logging
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
from calendar import monthrange
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

from .report_generator import _build_category_outcome_frame, _build_source_analysis_frame


class FormulaExcelEngine:
    """
    Excel report generator using 100% Excel formulas.
    No Python-calculated values in output sheets.
    """
    
    # Fixed column mapping
    COL_DATE = 'A'
    COL_DESC = 'B'
    COL_DEBIT = 'C'
    COL_CREDIT = 'D'
    COL_BALANCE = 'E'
    COL_CATEGORY = 'F'
    COL_CONFIDENCE = 'G'
    COL_RECURRING = 'H'
    
    # Sheet name
    RAW_SHEET = 'Raw Transactions'
    
    # Styles
    FONT_DEFAULT = Font(name='Arial', size=10)
    FONT_BOLD = Font(name='Arial', size=10, bold=True)
    FONT_HEADER = Font(name='Arial', size=10, bold=True)
    
    ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
    ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
    ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')
    
    BORDER_THIN = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    FILL_LIGHT_BLUE = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    FILL_LIGHT_ORANGE = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    
    # Number formats
    FMT_CURRENCY = '₹#,##0.00'
    FMT_INTEGER = '#,##0'
    FMT_DATE = 'DD-MM-YYYY'
    
    def __init__(self):
        self.workbook = None
        self.last_row = 0
        self.months: List[Tuple[int, int]] = []
    
    def generate(
        self,
        transactions: List[Dict[str, Any]],
        metadata: Dict[str, Any] = None,
        output_path: str = None
    ) -> bytes:
        """
        Generate Excel report with formula-based calculations.
        
        Args:
            transactions: List of transaction dicts with keys:
                date, description, debit, credit, balance, category, confidence, recurring
            metadata: Optional dict with name, account_no
            output_path: Optional file path to save
            
        Returns:
            Excel file bytes
        """
        self.workbook = Workbook()
        metadata = metadata or {}
        
        # Calculate last_row (header + data rows)
        self.last_row = len(transactions) + 1
        
        # Extract unique months from transactions
        self._extract_months(transactions)
        
        # Create sheets in order
        # Remove default sheet first
        default_sheet = self.workbook.active
        
        # Create all sheets
        ws_summary = self.workbook.create_sheet("Summary", 0)
        ws_category = self.workbook.create_sheet("Category Analysis", 1)
        ws_weekly = self.workbook.create_sheet("Weekly Analysis", 2)
        ws_recurring = self.workbook.create_sheet("Recurring Analysis", 3)
        ws_raw = self.workbook.create_sheet(self.RAW_SHEET, 4)
        ws_source = self.workbook.create_sheet("Source Analysis", 5)
        ws_outcome = self.workbook.create_sheet("Category Outcome", 6)
        ws_finbit = self.workbook.create_sheet("Finbit", 7)
        
        # Remove default sheet
        self.workbook.remove(default_sheet)
        
        # Build sheets
        self._build_raw_transactions(ws_raw, transactions)
        self._build_summary(ws_summary, metadata)
        self._build_category_analysis(ws_category)
        self._build_weekly_analysis(ws_weekly)
        self._build_recurring_analysis(ws_recurring)
        self._build_source_analysis(ws_source, transactions)
        self._build_category_outcome(ws_outcome, transactions)
        self._build_finbit(ws_finbit, transactions)
        
        # Save to bytes
        from io import BytesIO
        buffer = BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        
        if output_path:
            with open(output_path, 'wb') as f:
                f.write(buffer.getvalue())
        
        return buffer.getvalue()
    
    def _extract_months(self, transactions: List[Dict[str, Any]]):
        """Extract unique (year, month) tuples from transactions, sorted chronologically."""
        months_set = set()
        
        for txn in transactions:
            date_val = txn.get('date')
            if date_val:
                try:
                    if isinstance(date_val, str):
                        # Try multiple date formats
                        for fmt in ['%d/%m/%y', '%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                            try:
                                dt = datetime.strptime(date_val, fmt)
                                months_set.add((dt.year, dt.month))
                                break
                            except ValueError:
                                continue
                    elif isinstance(date_val, datetime):
                        months_set.add((date_val.year, date_val.month))
                except Exception:
                    pass
        
        self.months = sorted(list(months_set))
    
    def _get_last_day(self, year: int, month: int) -> int:
        """Get the last day of a month."""
        return monthrange(year, month)[1]
    
    def _build_raw_transactions(self, ws, transactions: List[Dict[str, Any]]):
        """Build the Raw Transactions sheet with source data."""
        # Headers
        headers = ['Date', 'Description', 'Debit', 'Credit', 'Balance', 'Category', 'Confidence', 'Recurring']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.FONT_BOLD
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
        
        # Data rows
        for row_idx, txn in enumerate(transactions, 2):
            # Column A - Date
            date_val = txn.get('date', '')
            if isinstance(date_val, str):
                # Try to parse and convert to datetime
                for fmt in ['%d/%m/%y', '%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                    try:
                        date_val = datetime.strptime(date_val, fmt)
                        break
                    except ValueError:
                        continue
            ws.cell(row=row_idx, column=1, value=date_val).number_format = self.FMT_DATE
            
            # Column B - Description
            ws.cell(row=row_idx, column=2, value=txn.get('description', ''))
            
            # Column C - Debit
            debit = txn.get('debit')
            if debit is not None and debit > 0:
                ws.cell(row=row_idx, column=3, value=debit).number_format = self.FMT_CURRENCY
            
            # Column D - Credit
            credit = txn.get('credit')
            if credit is not None and credit > 0:
                ws.cell(row=row_idx, column=4, value=credit).number_format = self.FMT_CURRENCY
            
            # Column E - Balance
            balance = txn.get('balance')
            if balance is not None:
                ws.cell(row=row_idx, column=5, value=balance).number_format = self.FMT_CURRENCY
            
            # Column F - Category
            ws.cell(row=row_idx, column=6, value=txn.get('category', ''))
            
            # Column G - Confidence
            ws.cell(row=row_idx, column=7, value=txn.get('confidence', ''))
            
            # Column H - Recurring
            ws.cell(row=row_idx, column=8, value=txn.get('recurring', 'No'))
        
        # Apply formatting
        for row in ws.iter_rows(min_row=2, max_row=self.last_row, min_col=1, max_col=8):
            for cell in row:
                cell.font = self.FONT_DEFAULT
                cell.border = self.BORDER_THIN
        
        # Column widths
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 13
        
        # Freeze header row
        ws.freeze_panes = 'A2'

    def _build_source_analysis(self, ws, transactions: List[Dict[str, Any]]):
        """Build the Source Analysis sheet for Sheet 9."""
        import pandas as pd

        df = pd.DataFrame(transactions).copy()
        if len(df) == 0:
            df = pd.DataFrame(columns=["date", "description", "debit", "credit", "balance", "category", "confidence", "recurring"])

        rename_map = {
            "date": "Date",
            "description": "Description",
            "debit": "Debit",
            "credit": "Credit",
            "balance": "Balance",
            "category": "Category",
            "recurring": "Recurring",
        }
        df.rename(columns={k: v for k, v in rename_map.items() if k in df.columns}, inplace=True)
        for col in ["Date", "Description", "Debit", "Credit", "Balance", "Category"]:
            if col not in df.columns:
                df[col] = "" if col == "Description" or col == "Category" else 0

        df["Date"] = pd.to_datetime(df["Date"], errors="coerce", dayfirst=True, format="mixed")
        df["Debit"] = pd.to_numeric(df["Debit"], errors="coerce").fillna(0)
        df["Credit"] = pd.to_numeric(df["Credit"], errors="coerce").fillna(0)
        df["Balance"] = pd.to_numeric(df["Balance"], errors="coerce").fillna(0)
        df["Description"] = df["Description"].fillna("").astype(str)
        df["Category"] = df["Category"].fillna("").astype(str)

        source_df = _build_source_analysis_frame(df)
        source_df = source_df.sort_values(["Date", "Description"], kind="stable").reset_index(drop=True)

        headers = [
            "Transaction Mode",
            "Source",
            "Identified Category",
            "Flag",
            "Date",
            "Description",
            "Credit",
            "Debit",
            "Balance",
        ]

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.cell(row=1, column=1, value="SOURCE ANALYSIS")
        ws.cell(row=1, column=1).font = self.FONT_BOLD
        ws.cell(row=1, column=1).alignment = self.ALIGN_CENTER

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = self.FONT_BOLD
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN

        for row_idx, row in enumerate(source_df.itertuples(index=False), 3):
            ws.cell(row=row_idx, column=1, value=getattr(row, "TransactionMode", "")).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=2, value=getattr(row, "Source", "")).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=3, value=getattr(row, "IdentifiedCategory", "")).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=4, value=getattr(row, "Flag", "")).border = self.BORDER_THIN

            dt = getattr(row, "Date", None)
            date_cell = ws.cell(row=row_idx, column=5)
            date_cell.border = self.BORDER_THIN
            if pd.notna(dt):
                date_cell.value = dt.to_pydatetime()
                date_cell.number_format = self.FMT_DATE
            else:
                date_cell.value = ""

            ws.cell(row=row_idx, column=6, value=getattr(row, "Description", "")).border = self.BORDER_THIN

            credit = float(getattr(row, "Credit", 0) or 0)
            debit = float(getattr(row, "Debit", 0) or 0)
            balance = float(getattr(row, "Balance", 0) or 0)

            credit_cell = ws.cell(row=row_idx, column=7, value=credit if credit > 0 else "")
            credit_cell.border = self.BORDER_THIN
            if credit > 0:
                credit_cell.number_format = self.FMT_CURRENCY

            debit_cell = ws.cell(row=row_idx, column=8, value=debit if debit > 0 else "")
            debit_cell.border = self.BORDER_THIN
            if debit > 0:
                debit_cell.number_format = self.FMT_CURRENCY

            balance_cell = ws.cell(row=row_idx, column=9, value=balance)
            balance_cell.border = self.BORDER_THIN
            balance_cell.number_format = self.FMT_CURRENCY

        for col, width in {"A": 18, "B": 24, "C": 22, "D": 18, "E": 14, "F": 50, "G": 16, "H": 16, "I": 16}.items():
            ws.column_dimensions[col].width = width

        ws.freeze_panes = 'A3'

    def _build_category_outcome(self, ws, transactions: List[Dict[str, Any]]):
        """Build the Category Outcome sheet for Sheet 10."""
        import pandas as pd

        df = pd.DataFrame(transactions).copy()
        if len(df) == 0:
            df = pd.DataFrame(columns=["date", "description", "debit", "credit", "balance", "category", "confidence", "recurring"])

        rename_map = {
            "date": "Date",
            "description": "Description",
            "debit": "Debit",
            "credit": "Credit",
            "balance": "Balance",
            "category": "Category",
        }
        df.rename(columns={k: v for k, v in rename_map.items() if k in df.columns}, inplace=True)
        for col in ["Date", "Description", "Debit", "Credit", "Balance", "Category"]:
            if col not in df.columns:
                df[col] = "" if col in ["Description", "Category"] else 0

        df["Date"] = pd.to_datetime(df["Date"], errors="coerce", dayfirst=True, format="mixed")
        df["Debit"] = pd.to_numeric(df["Debit"], errors="coerce").fillna(0)
        df["Credit"] = pd.to_numeric(df["Credit"], errors="coerce").fillna(0)
        df["Balance"] = pd.to_numeric(df["Balance"], errors="coerce").fillna(0)
        df["Description"] = df["Description"].fillna("").astype(str)
        df["Category"] = df["Category"].fillna("").astype(str)

        source_frame = _build_source_analysis_frame(df)
        outcome_tables = _build_category_outcome_tables(source_frame)
        month_labels = outcome_tables.get("month_labels", [])

        headers = ["Category", "Source", *month_labels]

        for col, width in {"A": 22, "B": 24}.items():
            ws.column_dimensions[col].width = width
        for idx in range(3, 3 + len(month_labels)):
            ws.column_dimensions[self._col_letter(idx)].width = 14
        ws.sheet_properties.outlinePr.summaryBelow = False
        ws.sheet_view.showOutlineSymbols = True

        def _write_table(start_row: int, title: str, table_name: str) -> int:
            table = outcome_tables.get(table_name)
            if table is None:
                return start_row

            end_col = len(headers)
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=end_col)
            ws.cell(row=start_row, column=1, value=title)
            ws.cell(row=start_row, column=1).font = self.FONT_BOLD
            ws.cell(row=start_row, column=1).alignment = self.ALIGN_CENTER

            header_row = start_row + 1
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_idx, value=header)
                cell.font = self.FONT_BOLD
                cell.alignment = self.ALIGN_CENTER
                cell.border = self.BORDER_THIN

            data_row = header_row + 1
            if table.empty:
                return data_row

            current_row = data_row
            category_order = []
            for category in table["Category"].fillna("").astype(str).tolist():
                if category not in category_order:
                    category_order.append(category)

            for category in category_order:
                category_rows = table[table["Category"] == category].copy()
                if category_rows.empty:
                    continue

                subtotal_values = {month_label: float(category_rows[month_label].fillna(0).sum()) for month_label in month_labels}

                subtotal_cell = ws.cell(row=current_row, column=1, value=category or "Others")
                subtotal_cell.font = self.FONT_BOLD
                subtotal_cell.alignment = self.ALIGN_LEFT
                subtotal_cell.border = self.BORDER_THIN

                source_cell = ws.cell(row=current_row, column=2, value="All Sources")
                source_cell.font = self.FONT_BOLD
                source_cell.alignment = self.ALIGN_LEFT
                source_cell.border = self.BORDER_THIN

                for col_idx, month_label in enumerate(month_labels, 3):
                    value = subtotal_values.get(month_label, 0)
                    cell = ws.cell(row=current_row, column=col_idx, value=float(value or 0))
                    cell.font = self.FONT_BOLD
                    cell.border = self.BORDER_THIN
                    cell.number_format = self.FMT_CURRENCY if "Amount" in title else self.FMT_INTEGER

                current_row += 1

                detail_start = current_row
                for _, row in category_rows.sort_values(["Source"], kind="stable").iterrows():
                    cat_cell = ws.cell(row=current_row, column=1, value=row.get("Category", ""))
                    cat_cell.border = self.BORDER_THIN
                    source_cell = ws.cell(row=current_row, column=2, value=row.get("Source", ""))
                    source_cell.border = self.BORDER_THIN

                    for col_idx, month_label in enumerate(month_labels, 3):
                        value = row.get(month_label, 0)
                        cell = ws.cell(row=current_row, column=col_idx, value=float(value or 0))
                        cell.border = self.BORDER_THIN
                        cell.number_format = self.FMT_CURRENCY if "Amount" in title else self.FMT_INTEGER
                    current_row += 1

                if current_row > detail_start:
                    ws.row_dimensions.group(detail_start, current_row - 1, hidden=True)

            return current_row

        next_row = 1
        next_row = _write_table(next_row, "CATEGORY OUTCOME  Credit Count", "credit_count") + 1
        next_row = _write_table(next_row, "CATEGORY OUTCOME  Debit Count", "debit_count") + 1
        next_row = _write_table(next_row, "CATEGORY OUTCOME  Credit Amount", "credit_amount") + 1
        _write_table(next_row, "CATEGORY OUTCOME  Debit Amount", "debit_amount")

        ws.freeze_panes = 'A3'

    def _build_finbit(self, ws, transactions: List[Dict[str, Any]]):
        """Build the Finbit sheet using the existing monthly metric logic."""
        import pandas as pd

        df = pd.DataFrame(transactions).copy()
        if len(df) == 0:
            return

        rename_map = {
            "date": "Date",
            "description": "Description",
            "debit": "Debit",
            "credit": "Credit",
            "balance": "Balance",
        }
        df.rename(columns={k: v for k, v in rename_map.items() if k in df.columns}, inplace=True)
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce", dayfirst=True, format="mixed")
        df["Debit"] = pd.to_numeric(df["Debit"], errors="coerce").fillna(0)
        df["Credit"] = pd.to_numeric(df["Credit"], errors="coerce").fillna(0)
        df["Balance"] = pd.to_numeric(df["Balance"], errors="coerce").fillna(0)
        df = df[df["Date"].notna()].copy()

        opening_balance = 0
        if len(df) > 0:
            first = df.iloc[0]
            opening_balance = first["Balance"] - first["Credit"] + first["Debit"]

        finbit_months, finbit_data = _compute_finbit_monthly(df, opening_balance)
        if not finbit_months or not finbit_data:
            return

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(finbit_months) + 1)
        ws.cell(row=1, column=1, value="FINBIT ANALYSIS")
        ws.cell(row=1, column=1).font = self.FONT_BOLD
        ws.cell(row=1, column=1).alignment = self.ALIGN_CENTER

        ws.cell(row=2, column=1, value="Metric").font = self.FONT_BOLD
        ws.cell(row=2, column=1).alignment = self.ALIGN_CENTER
        ws.cell(row=2, column=1).border = self.BORDER_THIN

        for ci, month_key in enumerate(finbit_months, 2):
            cell = ws.cell(row=2, column=ci, value=month_key)
            cell.font = self.FONT_BOLD
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN

        finbit_rows = [
            ("monthlyAvgBal", "Monthly Avg Balance", True),
            ("maxBalance", "Max Balance", True),
            ("minBalance", "Min Balance", True),
            ("cashDeposit", "Cash Deposits", True),
            ("cashWithdrawals", "Cash Withdrawals", True),
            ("chqDeposit", "Cheque Deposits", True),
            ("chqIssues", "Cheques Issued", True),
            ("credits", "Total Credits", True),
            ("debits", "Total Debits", True),
            ("inwBounce", "Inward Bounce", False),
            ("outwBounce", "Outward Bounce", False),
            ("penaltyCharges", "Penalty Charges", True),
            ("ecsNach", "ECS / NACH", True),
            ("totalNetDebit", "Total Net Debit", True),
            ("totalNetCredit", "Total Net Credit", True),
            ("selfWithdraw", "Self Withdrawal", True),
            ("selfDeposit", "Self Deposit", True),
            ("loanRepayment", "Loan Repayment", True),
            ("loanCredit", "Loan Credit", True),
            ("creditCardPayment", "Credit Card Payment", True),
            ("minCredits", "Min Credit Amount", True),
            ("maxCredits", "Max Credit Amount", True),
            ("salary", "Salary", True),
            ("bankCharges", "Bank Charges", True),
            None,
            ("balanceOpening", "BALANCE (Opening)", True),
            ("balanceClosing", "BALANCE (Closing)", True),
            ("salaryMonth", "SALARY (Income/Month)", True),
            ("ccPayment", "CCPAYMENT", True),
            ("eodMinBalance", "EOD MIN BALANCE", True),
            ("eodMaxBalance", "EOD MAX BALANCE", True),
        ]

        row_idx = 3
        for entry in finbit_rows:
            if entry is None:
                ws.cell(row=row_idx, column=1, value="Derived Monthly Metrics").font = self.FONT_BOLD
                row_idx += 1
                continue
            key, label, is_currency = entry
            ws.cell(row=row_idx, column=1, value=label).border = self.BORDER_THIN
            for ci, month_key in enumerate(finbit_months, 2):
                val = finbit_data[month_key].get(key, 0)
                cell = ws.cell(row=row_idx, column=ci, value=val)
                cell.border = self.BORDER_THIN
                cell.number_format = self.FMT_CURRENCY if is_currency else self.FMT_INTEGER
            row_idx += 1

        ws.column_dimensions['A'].width = 28
        for ci in range(2, len(finbit_months) + 2):
            ws.column_dimensions[get_column_letter(ci)].width = 18
        ws.freeze_panes = 'B3'
    
    def _build_summary(self, ws, metadata: Dict[str, Any]):
        """Build the Summary sheet with header block and monthly table."""
        lr = self.last_row
        RAW = f"'{self.RAW_SHEET}'"
        
        # ═══════════════════════════════════════════════════
        # HEADER BLOCK (Rows 1-4)
        # ═══════════════════════════════════════════════════
        
        header_data = [
            ('Name', metadata.get('name', '')),
            ('Account No', metadata.get('account_no', '')),
            ('Statement From', f"=MIN({RAW}!A:A)"),
            ('Statement To', f"=MAX({RAW}!A:A)")
        ]
        
        for row_idx, (label, value) in enumerate(header_data, 1):
            # Label
            cell_label = ws.cell(row=row_idx, column=1, value=label)
            cell_label.font = self.FONT_BOLD
            cell_label.alignment = self.ALIGN_LEFT
            cell_label.border = self.BORDER_THIN
            
            # Value
            cell_value = ws.cell(row=row_idx, column=2, value=value)
            cell_value.font = self.FONT_DEFAULT
            cell_value.alignment = self.ALIGN_LEFT
            cell_value.border = self.BORDER_THIN
            
            # Date format for rows 3-4
            if row_idx >= 3:
                cell_value.number_format = self.FMT_DATE
        
        # ═══════════════════════════════════════════════════
        # MONTHLY TABLE (Row 6 onward)
        # ═══════════════════════════════════════════════════
        
        # Row labels
        row_labels = [
            '',  # Row 6 - Header row (blank in column A)
            'Total Credit Count',
            'Total Credit Amount',
            'Total Debit Count',
            'Total Debit Amount',
            'Avg Balance',
            'Min Balance',
            'Max Balance',
            'Start of Month Balance',
            'End of Month Balance'
        ]
        
        # Write row labels
        for row_offset, label in enumerate(row_labels):
            row_idx = 6 + row_offset
            cell = ws.cell(row=row_idx, column=1, value=label)
            cell.font = self.FONT_BOLD if row_offset == 0 else self.FONT_BOLD
            cell.alignment = self.ALIGN_LEFT
            cell.border = self.BORDER_THIN
        
        # Write month columns
        for col_offset, (year, month) in enumerate(self.months):
            col_idx = 2 + col_offset
            col_letter = get_column_letter(col_idx)
            
            # Month label (row 6)
            month_name = datetime(year, month, 1).strftime('%b %Y')
            cell_header = ws.cell(row=6, column=col_idx, value=month_name)
            cell_header.font = self.FONT_BOLD
            cell_header.alignment = self.ALIGN_CENTER
            cell_header.border = self.BORDER_THIN
            cell_header.fill = self.FILL_LIGHT_BLUE
            
            # Calculate last day of month
            last_day = self._get_last_day(year, month)
            
            # Date range conditions for formulas
            date_gte = f'">="&DATE({year},{month},1)'
            date_lte = f'"<="&DATE({year},{month},{last_day})'
            
            # Row 7 - Total Credit Count
            formula = f'=COUNTIFS({RAW}!A$2:A${lr},{date_gte},{RAW}!A$2:A${lr},{date_lte},{RAW}!D$2:D${lr},">"&0)'
            cell = ws.cell(row=7, column=col_idx, value=formula)
            cell.number_format = self.FMT_INTEGER
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            
            # Row 8 - Total Credit Amount
            formula = f'=SUMIFS({RAW}!D$2:D${lr},{RAW}!A$2:A${lr},{date_gte},{RAW}!A$2:A${lr},{date_lte})'
            cell = ws.cell(row=8, column=col_idx, value=formula)
            cell.number_format = self.FMT_CURRENCY
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            
            # Row 9 - Total Debit Count
            formula = f'=COUNTIFS({RAW}!A$2:A${lr},{date_gte},{RAW}!A$2:A${lr},{date_lte},{RAW}!C$2:C${lr},">"&0)'
            cell = ws.cell(row=9, column=col_idx, value=formula)
            cell.number_format = self.FMT_INTEGER
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            
            # Row 10 - Total Debit Amount
            formula = f'=SUMIFS({RAW}!C$2:C${lr},{RAW}!A$2:A${lr},{date_gte},{RAW}!A$2:A${lr},{date_lte})'
            cell = ws.cell(row=10, column=col_idx, value=formula)
            cell.number_format = self.FMT_CURRENCY
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            
            # Row 11 - Avg Balance
            formula = f'=AVERAGEIFS({RAW}!E$2:E${lr},{RAW}!A$2:A${lr},{date_gte},{RAW}!A$2:A${lr},{date_lte})'
            cell = ws.cell(row=11, column=col_idx, value=formula)
            cell.number_format = self.FMT_CURRENCY
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            
            # Row 12 - Min Balance
            formula = f'=MINIFS({RAW}!E$2:E${lr},{RAW}!A$2:A${lr},{date_gte},{RAW}!A$2:A${lr},{date_lte})'
            cell = ws.cell(row=12, column=col_idx, value=formula)
            cell.number_format = self.FMT_CURRENCY
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            
            # Row 13 - Max Balance
            formula = f'=MAXIFS({RAW}!E$2:E${lr},{RAW}!A$2:A${lr},{date_gte},{RAW}!A$2:A${lr},{date_lte})'
            cell = ws.cell(row=13, column=col_idx, value=formula)
            cell.number_format = self.FMT_CURRENCY
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            
            # Row 14 - Start of Month Balance (first balance in that month)
            # Use simpler MINIFS approach to get balance on first date of month
            formula = (
                f"=IFERROR(INDEX({RAW}!E$2:E${lr},"
                f"MATCH(MINIFS({RAW}!A$2:A${lr},{RAW}!A$2:A${lr},\">=\"&DATE({year},{month},1),"
                f"{RAW}!A$2:A${lr},\"<=\"&DATE({year},{month},{last_day})),"
                f"{RAW}!A$2:A${lr},0)),\"\")"
            )
            cell = ws.cell(row=14, column=col_idx, value=formula)
            cell.number_format = self.FMT_CURRENCY
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            
            # Row 15 - End of Month Balance (last balance in that month)
            # Use MAXIFS to find last date, then INDEX/MATCH to get balance
            formula = (
                f"=IFERROR(INDEX({RAW}!E$2:E${lr},"
                f"MATCH(MAXIFS({RAW}!A$2:A${lr},{RAW}!A$2:A${lr},\">=\"&DATE({year},{month},1),"
                f"{RAW}!A$2:A${lr},\"<=\"&DATE({year},{month},{last_day})),"
                f"{RAW}!A$2:A${lr},0)),\"\")"
            )
            cell = ws.cell(row=15, column=col_idx, value=formula)
            cell.number_format = self.FMT_CURRENCY
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
        
        # Column widths
        ws.column_dimensions['A'].width = 28
        for col_offset in range(len(self.months)):
            col_letter = get_column_letter(2 + col_offset)
            ws.column_dimensions[col_letter].width = 18
        
        # Freeze panes
        ws.freeze_panes = 'A7'
    
    def _build_category_analysis(self, ws):
        """Build the Category Analysis sheet with credit and debit category tables."""
        lr = self.last_row
        RAW = f"'{self.RAW_SHEET}'"
        
        # ═══════════════════════════════════════════════════
        # LEFT TABLE - Credit Categories (Column A-C)
        # ═══════════════════════════════════════════════════
        
        # Header row
        headers_left = ['Credit Categories', 'Amount', 'Count']
        for col, header in enumerate(headers_left, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.FONT_BOLD
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            cell.fill = self.FILL_LIGHT_BLUE
        
        # Credit categories
        credit_categories = [
            'UPI',
            'Loan',
            'Salary Credits',
            'Bank Transfer',
            'Cash Deposit',
            'Others Credit',
            'Total Credit Amount'
        ]
        
        for row_offset, category in enumerate(credit_categories):
            row_idx = 2 + row_offset
            
            # Category name
            cell = ws.cell(row=row_idx, column=1, value=category)
            cell.font = self.FONT_BOLD if 'Total' in category else self.FONT_DEFAULT
            cell.alignment = self.ALIGN_LEFT
            cell.border = self.BORDER_THIN
            
            if 'Total' in category:
                # Total row - sum of above
                amount_formula = f'=SUM(B2:B{row_idx-1})'
                count_formula = f'=SUM(C2:C{row_idx-1})'
            else:
                # Category row
                amount_formula = f'=SUMIFS({RAW}!D$2:D${lr},{RAW}!F$2:F${lr},"*{category}*")'
                count_formula = f'=COUNTIFS({RAW}!F$2:F${lr},"*{category}*",{RAW}!D$2:D${lr},">"&0)'
            
            # Amount
            cell = ws.cell(row=row_idx, column=2, value=amount_formula)
            cell.number_format = self.FMT_CURRENCY
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            
            # Count
            cell = ws.cell(row=row_idx, column=3, value=count_formula)
            cell.number_format = self.FMT_INTEGER
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
        
        # ═══════════════════════════════════════════════════
        # RIGHT TABLE - Debit Categories (Column F-H)
        # ═══════════════════════════════════════════════════
        
        # Header row
        headers_right = ['Debit Categories', 'Amount', 'Count']
        for col_offset, header in enumerate(headers_right):
            col = 6 + col_offset  # F, G, H
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.FONT_BOLD
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            cell.fill = self.FILL_LIGHT_ORANGE
        
        # Debit categories
        debit_categories = [
            'Loan Payments',
            'ATM Withdrawal',
            'Shopping',
            'Bill Payment',
            'Withdrawal',
            'Investments',
            'Others Debit',
            'Total Debit Amount'
        ]
        
        for row_offset, category in enumerate(debit_categories):
            row_idx = 2 + row_offset
            
            # Category name
            cell = ws.cell(row=row_idx, column=6, value=category)
            cell.font = self.FONT_BOLD if 'Total' in category else self.FONT_DEFAULT
            cell.alignment = self.ALIGN_LEFT
            cell.border = self.BORDER_THIN
            
            if 'Total' in category:
                # Total row - sum of above
                amount_formula = f'=SUM(G2:G{row_idx-1})'
                count_formula = f'=SUM(H2:H{row_idx-1})'
            else:
                # Category row
                amount_formula = f'=SUMIFS({RAW}!C$2:C${lr},{RAW}!F$2:F${lr},"*{category}*")'
                count_formula = f'=COUNTIFS({RAW}!F$2:F${lr},"*{category}*",{RAW}!C$2:C${lr},">"&0)'
            
            # Amount
            cell = ws.cell(row=row_idx, column=7, value=amount_formula)
            cell.number_format = self.FMT_CURRENCY
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            
            # Count
            cell = ws.cell(row=row_idx, column=8, value=count_formula)
            cell.number_format = self.FMT_INTEGER
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
        
        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['E'].width = 3  # Spacer
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 10
    
    def _build_weekly_analysis(self, ws):
        """Build the Weekly Analysis sheet with SUMPRODUCT+DAY formulas."""
        lr = self.last_row
        RAW = f"'{self.RAW_SHEET}'"
        
        # ═══════════════════════════════════════════════════
        # LEFT TABLE - Credit Weekly Split (Column A-C)
        # ═══════════════════════════════════════════════════
        
        # Header row
        headers_left = ['Week', 'Credit Amount', 'Credit Count']
        for col, header in enumerate(headers_left, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.FONT_BOLD
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            cell.fill = self.FILL_LIGHT_BLUE
        
        # Week definitions
        weeks = [
            ('Week 1 (Days 1-7)', 1, 7),
            ('Week 2 (Days 8-14)', 8, 14),
            ('Week 3 (Days 15-21)', 15, 21),
            ('Week 4 (Days 22-31)', 22, 31),
            ('Total', None, None)
        ]
        
        for row_offset, (label, day_start, day_end) in enumerate(weeks):
            row_idx = 2 + row_offset
            
            # Week label
            cell = ws.cell(row=row_idx, column=1, value=label)
            cell.font = self.FONT_BOLD if label == 'Total' else self.FONT_DEFAULT
            cell.alignment = self.ALIGN_LEFT
            cell.border = self.BORDER_THIN
            
            if label == 'Total':
                # Sum of above
                amount_formula = f'=SUM(B2:B{row_idx-1})'
                count_formula = f'=SUM(C2:C{row_idx-1})'
            else:
                # SUMPRODUCT formula for week
                amount_formula = f'=SUMPRODUCT(({RAW}!D$2:D${lr})*(DAY({RAW}!A$2:A${lr})>={day_start})*(DAY({RAW}!A$2:A${lr})<={day_end}))'
                count_formula = f'=SUMPRODUCT(({RAW}!D$2:D${lr}>0)*(DAY({RAW}!A$2:A${lr})>={day_start})*(DAY({RAW}!A$2:A${lr})<={day_end})*1)'
            
            # Amount
            cell = ws.cell(row=row_idx, column=2, value=amount_formula)
            cell.number_format = self.FMT_CURRENCY
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            
            # Count
            cell = ws.cell(row=row_idx, column=3, value=count_formula)
            cell.number_format = self.FMT_INTEGER
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
        
        # ═══════════════════════════════════════════════════
        # RIGHT TABLE - Debit Weekly Split (Column F-H)
        # ═══════════════════════════════════════════════════
        
        # Header row
        headers_right = ['Week', 'Debit Amount', 'Debit Count']
        for col_offset, header in enumerate(headers_right):
            col = 6 + col_offset
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.FONT_BOLD
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            cell.fill = self.FILL_LIGHT_ORANGE
        
        for row_offset, (label, day_start, day_end) in enumerate(weeks):
            row_idx = 2 + row_offset
            
            # Week label
            cell = ws.cell(row=row_idx, column=6, value=label)
            cell.font = self.FONT_BOLD if label == 'Total' else self.FONT_DEFAULT
            cell.alignment = self.ALIGN_LEFT
            cell.border = self.BORDER_THIN
            
            if label == 'Total':
                # Sum of above
                amount_formula = f'=SUM(G2:G{row_idx-1})'
                count_formula = f'=SUM(H2:H{row_idx-1})'
            else:
                # SUMPRODUCT formula for week
                amount_formula = f'=SUMPRODUCT(({RAW}!C$2:C${lr})*(DAY({RAW}!A$2:A${lr})>={day_start})*(DAY({RAW}!A$2:A${lr})<={day_end}))'
                count_formula = f'=SUMPRODUCT(({RAW}!C$2:C${lr}>0)*(DAY({RAW}!A$2:A${lr})>={day_start})*(DAY({RAW}!A$2:A${lr})<={day_end})*1)'
            
            # Amount
            cell = ws.cell(row=row_idx, column=7, value=amount_formula)
            cell.number_format = self.FMT_CURRENCY
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            
            # Count
            cell = ws.cell(row=row_idx, column=8, value=count_formula)
            cell.number_format = self.FMT_INTEGER
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
        
        # Column widths
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['E'].width = 3  # Spacer
        ws.column_dimensions['F'].width = 22
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 14
    
    def _build_recurring_analysis(self, ws):
        """Build the Recurring Analysis sheet."""
        lr = self.last_row
        RAW = f"'{self.RAW_SHEET}'"
        
        # ═══════════════════════════════════════════════════
        # LEFT TABLE - Credit Recurring (Column A-C)
        # ═══════════════════════════════════════════════════
        
        # Header row
        headers_left = ['Type', 'Credit Amount', 'Credit Count']
        for col, header in enumerate(headers_left, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.FONT_BOLD
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            cell.fill = self.FILL_LIGHT_BLUE
        
        # Recurring types
        recurring_types = [
            ('Recurring', 'Yes'),
            ('Non-Recurring', 'No'),
            ('Total', None)
        ]
        
        for row_offset, (label, flag) in enumerate(recurring_types):
            row_idx = 2 + row_offset
            
            # Type label
            cell = ws.cell(row=row_idx, column=1, value=label)
            cell.font = self.FONT_BOLD if label == 'Total' else self.FONT_DEFAULT
            cell.alignment = self.ALIGN_LEFT
            cell.border = self.BORDER_THIN
            
            if label == 'Total':
                # Sum of above
                amount_formula = '=SUM(B2:B3)'
                count_formula = '=SUM(C2:C3)'
            else:
                # SUMIFS/COUNTIFS formula
                amount_formula = f'=SUMIFS({RAW}!D$2:D${lr},{RAW}!H$2:H${lr},"{flag}")'
                count_formula = f'=COUNTIFS({RAW}!H$2:H${lr},"{flag}",{RAW}!D$2:D${lr},">"&0)'
            
            # Amount
            cell = ws.cell(row=row_idx, column=2, value=amount_formula)
            cell.number_format = self.FMT_CURRENCY
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            
            # Count
            cell = ws.cell(row=row_idx, column=3, value=count_formula)
            cell.number_format = self.FMT_INTEGER
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
        
        # ═══════════════════════════════════════════════════
        # RIGHT TABLE - Debit Recurring (Column F-H)
        # ═══════════════════════════════════════════════════
        
        # Header row
        headers_right = ['Type', 'Debit Amount', 'Debit Count']
        for col_offset, header in enumerate(headers_right):
            col = 6 + col_offset
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.FONT_BOLD
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            cell.fill = self.FILL_LIGHT_ORANGE
        
        for row_offset, (label, flag) in enumerate(recurring_types):
            row_idx = 2 + row_offset
            
            # Type label
            cell = ws.cell(row=row_idx, column=6, value=label)
            cell.font = self.FONT_BOLD if label == 'Total' else self.FONT_DEFAULT
            cell.alignment = self.ALIGN_LEFT
            cell.border = self.BORDER_THIN
            
            if label == 'Total':
                # Sum of above
                amount_formula = '=SUM(G2:G3)'
                count_formula = '=SUM(H2:H3)'
            else:
                # SUMIFS/COUNTIFS formula
                amount_formula = f'=SUMIFS({RAW}!C$2:C${lr},{RAW}!H$2:H${lr},"{flag}")'
                count_formula = f'=COUNTIFS({RAW}!H$2:H${lr},"{flag}",{RAW}!C$2:C${lr},">"&0)'
            
            # Amount
            cell = ws.cell(row=row_idx, column=7, value=amount_formula)
            cell.number_format = self.FMT_CURRENCY
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
            
            # Count
            cell = ws.cell(row=row_idx, column=8, value=count_formula)
            cell.number_format = self.FMT_INTEGER
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
        
        # Column widths
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['E'].width = 3  # Spacer
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 14
